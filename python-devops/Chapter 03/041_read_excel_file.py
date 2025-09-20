# Script: 041_read_excel_file.py

# Importing the load_workbook function from openpyxl to load existing Excel files
from openpyxl import load_workbook

# Load an existing workbook
wb = load_workbook("example.xlsx")
sheet = wb.active

# Read data from a specific cell (A1)
print(sheet['A1'].value)

# Read data from multiple cells (first two rows and first two columns)
for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=2):
    for cell in row:
        print(cell.value)
